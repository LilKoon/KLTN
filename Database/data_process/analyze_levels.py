"""
Phân tích tần suất thực để điều chỉnh ngưỡng Level cho 3 mức hợp lý.
"""
import openpyxl
import re
import statistics
from wordfreq import word_frequency
from collections import Counter

EXCEL_PATH = r'c:\Bun\KLTN_main\Database\listen_data\Dapan_Listenning.xlsx'

SIMPLE_STOPWORDS = {
    'a', 'an', 'the', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
    'i', 'you', 'he', 'she', 'it', 'we', 'they', 'me', 'him', 'her', 'us', 'them',
    'my', 'your', 'his', 'its', 'our', 'their', 'this', 'that', 'these', 'those',
    'to', 'of', 'in', 'on', 'at', 'by', 'for', 'with', 'from', 'and', 'or', 'but',
    'not', 'no', 'so', 'if', 'do', 'did', 'does', 'have', 'has', 'had',
    'will', 'would', 'could', 'should', 'can', 'may', 'might', 'shall',
    'up', 'out', 'as', 'into', 'about', 'than', 'its', 'all', 'any', 'just'
}

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active
headers = [cell.value for cell in ws[1]]
col_idx = {h: i for i, h in enumerate(headers)}

median_freqs = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    audio = str(row[col_idx['AudioFile']] or '').strip()
    transcript = str(row[col_idx['Transcript']] or '').strip()
    opt_a = str(row[col_idx['Option_A']] or '').strip()
    opt_b = str(row[col_idx['Option_B']] or '').strip()
    opt_c = str(row[col_idx['Option_C']] or '').strip()
    
    if not audio or audio.lower() == 'none':
        continue
    
    text = transcript if len(transcript) > 30 else f"{transcript} {opt_a} {opt_b} {opt_c}"
    words = re.findall(r'[a-zA-Z]+', text.lower())
    content_words = [w for w in words if w not in SIMPLE_STOPWORDS and len(w) > 2]
    
    if not content_words:
        continue
    
    freqs = [word_frequency(w, 'en') for w in content_words]
    mf = statistics.median(freqs)
    median_freqs.append(mf)

print(f"Total transcripts analyzed: {len(median_freqs)}")
print(f"Min freq: {min(median_freqs):.2e}")
print(f"Max freq: {max(median_freqs):.2e}")
print(f"Overall median: {statistics.median(median_freqs):.2e}")
print(f"Mean: {statistics.mean(median_freqs):.2e}")

# Tính percentiles để chia 3 mức đều
sorted_freqs = sorted(median_freqs)
n = len(sorted_freqs)
p33 = sorted_freqs[n//3]
p66 = sorted_freqs[2*n//3]
print(f"\nPercentile 33%: {p33:.2e}")
print(f"Percentile 66%: {p66:.2e}")

print("\nSuggested thresholds:")
print(f"  Level 1 (Beginner):     freq >= {p66:.2e}")
print(f"  Level 2 (Intermediate): {p33:.2e} <= freq < {p66:.2e}")
print(f"  Level 3 (Advanced):     freq < {p33:.2e}")

# Preview phân phối với ngưỡng mới
levels = []
for f in median_freqs:
    if f >= p66:
        levels.append(1)
    elif f >= p33:
        levels.append(2)
    else:
        levels.append(3)
cnt = Counter(levels)
print(f"\nDistribution with new thresholds: {dict(sorted(cnt.items()))}")
